import json
import logging
import os
from collections.abc import AsyncIterator
from pathlib import Path
from typing import Any, Optional

from fastapi import APIRouter, FastAPI, File, Form, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field

from backend.config import UPLOAD_DIR
from backend.rag.service import default_rag_service
from backend.ai.agent_loop import AgentResponse, ReActAgentEngine, react_agent_engine
from backend.ai.model_types import ComplexityType
from backend.ai.registry import ModelRegistry, model_registry
from backend.ai.router import DynamicRouter, model_router
from backend.ai.tools import (
    ToolRegistry,
    calculate_expression,
    run_python_sandbox,
    search_documents,
    tool_registry,
)
from backend.ai.workflows import (
    run_coding_sandbox_workflow,
    run_document_approval_workflow,
    run_document_qa_workflow,
    run_multimodal_analysis_workflow,
    stream_autoroute_workflow,
    stream_coding_sandbox_workflow,
    stream_document_approval_workflow,
    stream_document_qa_workflow,
    stream_multimodal_analysis_workflow,
)
from backend.ai.workspace_manager import WorkspaceSecurityError, workspace_manager

logger = logging.getLogger(__name__)

# Router for Module 3 API
router = APIRouter(prefix="/api/v1", tags=["Agent & RAG Pipeline"])


# ==========================================
# Pydantic Request & Response Models
# ==========================================

class RAGChatRequest(BaseModel):
    question: str = Field(..., description="The user question to answer from indexed documents")
    n_results: int = Field(default=3, ge=1, le=20, description="Number of evidence chunks to retrieve")
    source: Optional[str] = Field(None, description="Optional document source/filename for strict source isolation")
    model_override: Optional[str] = Field(None, description="Explicit model name override (e.g. 'qwen3-4b')")


class RAGChatResponse(BaseModel):
    answer: str
    sources: list[dict[str, Any]] = Field(default_factory=list)
    contexts: list[dict[str, Any]] = Field(default_factory=list)


class DocumentUploadResponse(BaseModel):
    status: str = "success"
    filename: str
    chunks_indexed: int
    vectors_stored: int
    message: str = "Document successfully extracted, chunked, embedded, and stored."


class ChatAgentRequest(BaseModel):
    message: str = Field(..., description="The user prompt or instruction")
    conversation_id: Optional[str] = Field(None, description="Unique conversation thread ID")
    media_paths: Optional[list[str]] = Field(default_factory=list, description="Paths to attached image or video files")
    history: Optional[list[dict[str, Any]]] = Field(default_factory=list, description="Previous conversation turn history")
    model_override: Optional[str] = Field(None, description="Explicit model name override (e.g. 'qwen3-8b')")
    complexity_override: Optional[ComplexityType] = Field(None, description="Explicit complexity level ('low', 'medium', 'high')")
    max_vram_gb: Optional[float] = Field(None, description="VRAM budget ceiling in GB")
    system_prompt: Optional[str] = Field(None, description="Custom system prompt override")
    stream: bool = Field(False, description="Enable real-time SSE event streaming")


class ChatAgentResponse(BaseModel):
    conversation_id: str
    sender: str = "assistant"
    model_used: str
    content: str
    citations: list[dict[str, Any]] = Field(default_factory=list)
    tool_calls: list[dict[str, Any]] = Field(default_factory=list)
    routing_decision: Optional[dict[str, Any]] = None
    execution_time_seconds: float
    turns_count: int


class DocumentSearchRequest(BaseModel):
    query: str
    top_k: int = 3
    department: Optional[str] = None


class CalculatorRequest(BaseModel):
    expression: str


class SandboxRequest(BaseModel):
    code: str
    timeout_seconds: int = 5


class DocumentApprovalRequest(BaseModel):
    document_path: str = Field(default="documents/inspection_report.pdf", description="Relative path in workspace")
    prompt: Optional[str] = Field(None, description="Optional custom instruction")
    output_filename: Optional[str] = Field(None, description="Custom output filename")


class DocumentQARequest(BaseModel):
    document_path: str = Field(..., description="Relative path or filename of the document in workspace")
    question: str = Field(..., description="User question to answer directly from the attached document")


class CodingSandboxWorkflowRequest(BaseModel):
    prompt: str = Field(default="Write a Python program to calculate pressure drop using the Darcy-Weisbach equation.", description="Coding problem")
    timeout_seconds: int = Field(default=10, ge=1, le=60)


class MultimodalWorkflowRequest(BaseModel):
    image_path: str = Field(default="documents/inspection_report.pdf", description="Relative path in workspace")
    prompt: Optional[str] = Field(None, description="Vision instruction")


class AutoRouteWorkflowRequest(BaseModel):
    query: str = Field(..., description="Task prompt or multimodal instruction to automatically route and execute")


# ==========================================
# Canonical Endpoints
# ==========================================

@router.post("/documents/upload", response_model=DocumentUploadResponse)
async def upload_document(
    file: UploadFile = File(...),
    department: Optional[str] = Form(None),
):
    """
    Canonical Document Ingestion Endpoint.
    Receives PDF/DOCX/ODT/TXT/Images -> PyMuPDF extraction + OCR fallback ->
    chunking -> metadata extraction -> 384-d embedding -> PostgreSQL/VectorStore.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Filename is required.")

    filename = Path(file.filename).name
    suffix = Path(filename).suffix.lower()

    supported = {
        ".pdf", ".docx", ".odt", ".txt", ".md",
        ".png", ".jpg", ".jpeg", ".tiff", ".bmp", ".webp",
    }

    if suffix not in supported:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{suffix}'. Supported formats: {', '.join(sorted(supported))}",
        )

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    upload_path = UPLOAD_DIR / filename
    try:
        upload_path.write_bytes(contents)
        result = default_rag_service.ingest(upload_path)

        return DocumentUploadResponse(
            status="success",
            filename=filename,
            chunks_indexed=result["chunks_indexed"],
            vectors_stored=result["vectors_stored"],
            message=f"Indexed {result['chunks_indexed']} chunks from {filename} into vector store.",
        )
    except Exception as exc:
        logger.error(f"Document ingestion error: {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Document processing failed: {exc}")


@router.post("/chat", response_model=RAGChatResponse)
async def rag_chat(request: RAGChatRequest):
    """
    Canonical Grounded RAG Chat Endpoint.
    Hybrid retrieval (semantic + exact metadata) -> VaultMind model router -> local Ollama -> Grounded answer with citations.
    """
    if not request.question.strip():
        raise HTTPException(status_code=400, detail="Question cannot be empty.")

    try:
        result = default_rag_service.ask(
            question=request.question,
            n_results=request.n_results,
            source=request.source,
        )

        sources = []
        for context in result.get("contexts", []):
            metadata = context.get("metadata", {})
            src = metadata.get("source")
            if not src:
                continue

            source_info = {
                "source": src,
                "page": metadata.get("page", 1),
                "chunk_id": metadata.get("chunk_id"),
                "ocr_used": metadata.get("ocr_used", False),
                "document_id": metadata.get("document_id"),
                "equipment_tag": metadata.get("equipment_tag"),
                "distance": context.get("distance", 0.0),
            }
            if source_info not in sources:
                sources.append(source_info)

        return RAGChatResponse(
            answer=result.get("answer", ""),
            sources=sources,
            contexts=result.get("contexts", []),
        )
    except Exception as exc:
        logger.error(f"RAG query failed: {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"RAG query failed: {exc}")


@router.post("/chat/agent", response_model=ChatAgentResponse)
async def chat_agent(request: ChatAgentRequest):
    """
    Main ReAct Cognitive Agent Endpoint.
    Dynamically routes to optimal model, executes multi-turn tool calling,
    and returns final structured response or SSE stream.
    """
    if request.stream:
        async def event_generator() -> AsyncIterator[str]:
            async for ev in react_agent_engine.stream_run(
                message=request.message,
                conversation_id=request.conversation_id,
                media_paths=request.media_paths,
                history=request.history,
                model_override=request.model_override,
                complexity_override=request.complexity_override,
                max_vram_gb=request.max_vram_gb,
                system_prompt=request.system_prompt,
            ):
                event_type = ev.get("event", "message")
                event_data = json.dumps(ev.get("data", {}), ensure_ascii=False)
                yield f"event: {event_type}\ndata: {event_data}\n\n"

        return StreamingResponse(
            event_generator(),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Accel-Buffering": "no",
            },
        )

    response = await react_agent_engine.run(
        message=request.message,
        conversation_id=request.conversation_id,
        media_paths=request.media_paths,
        history=request.history,
        model_override=request.model_override,
        complexity_override=request.complexity_override,
        max_vram_gb=request.max_vram_gb,
        system_prompt=request.system_prompt,
    )

    return ChatAgentResponse(
        conversation_id=response.conversation_id,
        sender=response.sender,
        model_used=response.model_used,
        content=response.content,
        citations=response.citations,
        tool_calls=response.tool_calls,
        routing_decision=response.routing_decision,
        execution_time_seconds=response.execution_time_seconds,
        turns_count=response.turns_count,
    )


@router.post("/chat/completions")
async def chat_completions(request: ChatAgentRequest):
    """Alias endpoint for compatibility with OpenAI/standard chat format."""
    return await chat_agent(request)


@router.get("/stats")
async def get_stats():
    """Corpus, vector store, and model statistics."""
    vector_store_backend = getattr(default_rag_service.vector_store, "backend_type", "postgresql")
    is_pg = getattr(default_rag_service.vector_store, "is_pgvector", True)
    return {
        "vector_store": vector_store_backend,
        "pgvector": is_pg,
        "vectors_stored": default_rag_service.count(),
        "models_registered": len(model_registry.list_models()),
        "tools_registered": len(tool_registry.list_tools()),
    }


@router.get("/models")
async def list_models():
    """List all registered LLM and Multimodal models with metadata."""
    models = model_registry.list_models()
    return {
        "count": len(models),
        "models": [
            {
                "name": m.name,
                "ollama_model": m.ollama_model,
                "modalities": list(m.modalities),
                "capabilities": list(m.capabilities),
                "max_complexity": m.max_complexity,
                "target_vram_gb": m.target_vram_gb,
                "tool_calling": m.tool_calling,
                "reasoning": m.reasoning,
            }
            for m in models
        ],
    }


@router.get("/tools")
async def list_tools():
    """List all available agent tools and schemas."""
    tools = tool_registry.list_tools()
    return {
        "count": len(tools),
        "tools": [
            {
                "name": t.name,
                "description": t.description,
                "required_skill": t.required_skill,
                "parameters": t.parameters,
            }
            for t in tools
        ],
    }


@router.post("/documents/search")
async def search_docs(request: DocumentSearchRequest):
    """Vector similarity retrieval against document repository."""
    return search_documents(
        query=request.query,
        top_k=request.top_k,
        department=request.department,
    )


@router.post("/calculator")
async def calc(request: CalculatorRequest):
    """Safe mathematical expression evaluator."""
    return calculate_expression(request.expression)


@router.post("/sandbox/execute")
async def run_sandbox(request: SandboxRequest):
    """Execute Python code in isolated sandbox subprocess."""
    return run_python_sandbox(
        code=request.code,
        timeout_seconds=request.timeout_seconds,
    )


@router.get("/health")
async def health_check():
    """System health check, RAG status, and local connectivity."""
    vector_store_backend = getattr(default_rag_service.vector_store, "backend_type", "postgresql")
    is_pg = getattr(default_rag_service.vector_store, "is_pgvector", True)
    return {
        "status": "healthy",
        "service": "VaultMind Sovereign AI",
        "air_gap_compliant": True,
        "rag_service": "online",
        "vector_store": vector_store_backend,
        "pgvector": is_pg,
        "vectors_stored": default_rag_service.count(),
        "models_registered": len(model_registry.list_models()),
        "tools_registered": len(tool_registry.list_tools()),
        "workspace_root": str(workspace_manager.root),
    }


# ==========================================
# Controlled Workspace & File Management APIs
# ==========================================

@router.get("/workspace/tree")
async def get_workspace_tree():
    """Get recursive folder structure of the controlled workspace."""
    return workspace_manager.get_tree()


@router.post("/workspace/upload")
async def upload_to_workspace(
    file: UploadFile = File(...),
    subdir: str = Form(default="input"),
):
    """
    Upload a file safely into a controlled workspace subdirectory (e.g. 'documents', 'input').
    Automatically triggers background RAG vector indexing if the file is a document.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="Filename required.")
    
    contents = await file.read()
    try:
        saved_path = workspace_manager.save_file(subdir=subdir, filename=file.filename, content=contents)
        rel_path = str(saved_path.relative_to(workspace_manager.root))
        
        # Auto-index into RAG vector store if placed in documents or input
        chunks_indexed = 0
        if saved_path.suffix.lower() in (".pdf", ".docx", ".txt", ".md", ".png", ".jpg", ".jpeg"):
            try:
                ingest_res = default_rag_service.ingest(saved_path)
                chunks_indexed = ingest_res.get("chunks_indexed", 0)
            except Exception as e:
                logger.warning(f"Auto-ingest warning for {saved_path.name}: {e}")

        return {
            "status": "success",
            "filename": saved_path.name,
            "relative_path": rel_path,
            "size_bytes": len(contents),
            "chunks_indexed": chunks_indexed,
            "message": f"Saved '{saved_path.name}' to workspace/{rel_path}",
        }
    except WorkspaceSecurityError as sec_err:
        raise HTTPException(status_code=403, detail=str(sec_err))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Upload failed: {exc}")


@router.get("/workspace/file")
async def download_workspace_file(path: str = Query(..., description="Relative file path in workspace")):
    """Safely retrieve or download a file from the controlled workspace.
    
    Checks workspace/output/ first (primary), then project-root output/ for backward compatibility.
    """
    from pathlib import Path as _Path
    _project_root = _Path(__file__).resolve().parent.parent.parent

    # Primary: workspace/output/<filename>
    try:
        safe_path = workspace_manager.resolve_safe_path(path)
    except WorkspaceSecurityError as sec_err:
        # Maybe path includes full absolute — try just the filename
        filename = _Path(path).name
        try:
            safe_path = workspace_manager.resolve_safe_path(f"output/{filename}")
        except Exception:
            raise HTTPException(status_code=403, detail=str(sec_err))

    # Fallback: project-root output/ directory
    if not safe_path.exists():
        fallback = _project_root / "output" / _Path(path).name
        if fallback.exists():
            safe_path = fallback

    if not safe_path.exists() or not safe_path.is_file():
        raise HTTPException(status_code=404, detail=f"File '{path}' not found in workspace.")

    media_type = "application/octet-stream"
    if safe_path.suffix.lower() == ".pdf":
        media_type = "application/pdf"
    elif safe_path.suffix.lower() == ".docx":
        media_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    elif safe_path.suffix.lower() in (".xlsx", ".xlsm"):
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif safe_path.suffix.lower() in (".png", ".jpg", ".jpeg", ".webp"):
        media_type = f"image/{safe_path.suffix.lower().lstrip('.')}"
    elif safe_path.suffix.lower() in (".txt", ".py", ".md", ".log", ".json", ".csv"):
        media_type = "text/plain"

    return FileResponse(
        path=str(safe_path),
        filename=safe_path.name,
        media_type=media_type,
    )


@router.get("/workspace/sheet/preview")
async def preview_workspace_sheet(path: str = Query(..., description="Relative Excel or CSV file path")):
    """Read an Excel (.xlsx) or CSV file and return structured headers, rows, and sheets for UI rendering."""
    try:
        safe_path = workspace_manager.resolve_safe_path(path)
        if not safe_path.exists() or not safe_path.is_file():
            raise HTTPException(status_code=404, detail=f"Sheet file '{path}' not found.")

        if safe_path.suffix.lower() in (".xlsx", ".xlsm"):
            import openpyxl
            wb = openpyxl.load_workbook(str(safe_path), data_only=True)
            sheet_names = wb.sheetnames
            ws = wb.active
            rows_data = []
            for row in ws.iter_rows(values_only=True):
                if any(v is not None for v in row):
                    rows_data.append([str(c) if c is not None else "" for c in row])

            headers = rows_data[0] if rows_data else []
            rows = rows_data[1:] if len(rows_data) > 1 else []
            return {
                "status": "success",
                "filename": safe_path.name,
                "sheets": sheet_names,
                "active_sheet": ws.title,
                "headers": headers,
                "rows": rows,
                "total_rows": len(rows),
                "total_cols": len(headers),
            }
        elif safe_path.suffix.lower() == ".csv":
            import csv
            rows_data = []
            with open(str(safe_path), "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                for r in reader:
                    if r:
                        rows_data.append(r)
            headers = rows_data[0] if rows_data else []
            rows = rows_data[1:] if len(rows_data) > 1 else []
            return {
                "status": "success",
                "filename": safe_path.name,
                "sheets": ["Sheet1"],
                "active_sheet": "Sheet1",
                "headers": headers,
                "rows": rows,
                "total_rows": len(rows),
                "total_cols": len(headers),
            }
        else:
            raise HTTPException(status_code=400, detail="File is not a supported spreadsheet format (.xlsx, .csv).")
    except WorkspaceSecurityError as sec_err:
        raise HTTPException(status_code=403, detail=str(sec_err))
    except Exception as exc:
        logger.error(f"Sheet preview error: {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to read sheet: {exc}")


@router.delete("/workspace/file")
async def delete_workspace_file(path: str = Query(..., description="Relative file path in workspace")):
    """Safely delete a file from the controlled workspace."""
    try:
        deleted = workspace_manager.delete_file(path)
        if not deleted:
            raise HTTPException(status_code=404, detail=f"File '{path}' not found.")
        return {"status": "success", "message": f"Deleted '{path}'"}
    except WorkspaceSecurityError as sec_err:
        raise HTTPException(status_code=403, detail=str(sec_err))


# ==========================================
# Specialized SIH Demonstration Workflows
# ==========================================

@router.post("/workspace/workflow/document-approval")
async def workflow_document_approval(request: DocumentApprovalRequest):
    """
    Flagship SIH Demo 1:
    Scanned Document -> OCR -> Extract Key Findings -> Reasoning Model -> Draft Approval Note -> Real DOCX artifact.
    """
    try:
        return await run_document_approval_workflow(
            document_rel_path=request.document_path,
            prompt=request.prompt,
            output_filename=request.output_filename,
        )
    except FileNotFoundError as fnf:
        raise HTTPException(status_code=404, detail=str(fnf))
    except WorkspaceSecurityError as sec_err:
        raise HTTPException(status_code=403, detail=str(sec_err))
    except Exception as exc:
        logger.error(f"Document approval workflow error: {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Workflow failed: {exc}")


@router.post("/workspace/workflow/document-approval/stream")
async def workflow_document_approval_stream(request: DocumentApprovalRequest):
    """Real-time SSE event stream for Document Approval Workflow."""
    async def event_generator() -> AsyncIterator[str]:
        async for ev in stream_document_approval_workflow(
            document_rel_path=request.document_path,
            prompt=request.prompt,
            output_filename=request.output_filename,
        ):
            event_type = ev.get("event", "message")
            event_data = json.dumps(ev.get("data", {}), ensure_ascii=False)
            yield f"event: {event_type}\ndata: {event_data}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
    )


@router.post("/workspace/workflow/document-qa")
async def workflow_document_qa(request: DocumentQARequest):
    """Direct QA on an attached document without searching external database."""
    try:
        return await run_document_qa_workflow(
            document_rel_path=request.document_path,
            question=request.question,
        )
    except Exception as exc:
        logger.error(f"Document QA workflow error: {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Document QA failed: {exc}")


@router.post("/workspace/workflow/document-qa/stream")
async def workflow_document_qa_stream(request: DocumentQARequest):
    """Real-time SSE event stream for direct Document QA Workflow."""
    async def event_generator() -> AsyncIterator[str]:
        async for ev in stream_document_qa_workflow(
            document_rel_path=request.document_path,
            question=request.question,
        ):
            event_type = ev.get("event", "message")
            event_data = json.dumps(ev.get("data", {}), ensure_ascii=False)
            yield f"event: {event_type}\ndata: {event_data}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
    )


@router.post("/workspace/workflow/coding-sandbox")
async def workflow_coding_sandbox(request: CodingSandboxWorkflowRequest):
    """
    Flagship SIH Demo 2:
    Coding Request -> Local Coding Model -> Generate Code -> Execute in Sandbox -> Verified Output & Tests.
    """
    try:
        return await run_coding_sandbox_workflow(
            prompt=request.prompt,
            timeout_seconds=request.timeout_seconds,
        )
    except Exception as exc:
        logger.error(f"Coding sandbox workflow error: {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Workflow failed: {exc}")


@router.post("/workspace/workflow/coding-sandbox/stream")
async def workflow_coding_sandbox_stream(request: CodingSandboxWorkflowRequest):
    """Real-time SSE event stream for Coding Sandbox Workflow."""
    async def event_generator() -> AsyncIterator[str]:
        async for ev in stream_coding_sandbox_workflow(
            prompt=request.prompt,
            timeout_seconds=request.timeout_seconds,
        ):
            event_type = ev.get("event", "message")
            event_data = json.dumps(ev.get("data", {}), ensure_ascii=False)
            yield f"event: {event_type}\ndata: {event_data}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
    )


@router.post("/workspace/workflow/multimodal")
async def workflow_multimodal(request: MultimodalWorkflowRequest):
    """
    Flagship SIH Demo 3:
    Image/Drawing Input -> Local Vision Model (Qwen3-VL) -> Structured Findings.
    """
    try:
        return await run_multimodal_analysis_workflow(
            image_rel_path=request.image_path,
            prompt=request.prompt,
        )
    except FileNotFoundError as fnf:
        raise HTTPException(status_code=404, detail=str(fnf))
    except WorkspaceSecurityError as sec_err:
        raise HTTPException(status_code=403, detail=str(sec_err))
    except Exception as exc:
        logger.error(f"Multimodal workflow error: {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Workflow failed: {exc}")


@router.post("/workspace/workflow/multimodal/stream")
async def workflow_multimodal_stream(request: MultimodalWorkflowRequest):
    """Real-time SSE event stream for Multimodal Vision Workflow."""
    async def event_generator() -> AsyncIterator[str]:
        async for ev in stream_multimodal_analysis_workflow(
            image_rel_path=request.image_path,
            prompt=request.prompt,
        ):
            event_type = ev.get("event", "message")
            event_data = json.dumps(ev.get("data", {}), ensure_ascii=False)
            yield f"event: {event_type}\ndata: {event_data}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
    )


@router.post("/workspace/workflow/autoroute/stream")
async def workflow_autoroute_stream(request: AutoRouteWorkflowRequest):
    """
    Flagship SIH Demo 4:
    Analyzes intent -> Determines workflow & model -> Executes real pipeline -> Streams real-time SSE.
    """
    async def event_generator() -> AsyncIterator[str]:
        async for ev in stream_autoroute_workflow(query=request.query):
            event_type = ev.get("event", "message")
            event_data = json.dumps(ev.get("data", {}), ensure_ascii=False)
            yield f"event: {event_type}\ndata: {event_data}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "Connection": "keep-alive", "X-Accel-Buffering": "no"},
    )



def create_app() -> FastAPI:
    """Create and configure FastAPI application."""
    app = FastAPI(
        title="VaultMind Sovereign AI - Backend API",
        description="Air-gapped, multimodal RAG agent backend powered by local Ollama models and PostgreSQL pgvector",
        version="1.0.0",
    )

    # Enable CORS for frontend development
    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )

    app.include_router(router)
    return app


# Default app instance
app = create_app()
